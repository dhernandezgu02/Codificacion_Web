import os
import flet as ft
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
from openai import OpenAI
import re
import time
from tqdm import tqdm
from config import openai_api_key_Codifiacion

# Configure OpenAI API
client = OpenAI(api_key=openai_api_key_Codifiacion)

# Agregar esta variable global al inicio del archivo (después de los imports)
PROCESS_STOPPED = False

# Variable global para rastrear celdas modificadas
MODIFIED_CELLS = set()



# Funciones principales del script original (copiadas directamente)
def load_files(responses_path, codes_path):
    responses_df = pd.read_excel(responses_path)
    codes_df = pd.read_excel(codes_path, sheet_name='Codificación')
    codes_df.columns = codes_df.columns.str.strip()
    return responses_df, codes_df

def select_columns(codes_df, question_column):
    codes_df[question_column] = codes_df[question_column].ffill()
    selected_questions = codes_df[[question_column, 'Label', 'Id campo', 'Cod']]
    return selected_questions

# Modificar la función request_openai para verificar directamente la variable global
def request_openai(messages, max_retries=5, stop_requested_check=None):
    global PROCESS_STOPPED
    
    for attempt in range(max_retries):
        # Verificar directamente la variable global
        if PROCESS_STOPPED or (stop_requested_check and stop_requested_check()):
            print("Solicitud a OpenAI cancelada - proceso detenido")
            return None
        
        try:
            response = client.chat.completions.create(
                model="gpt-4o",
                messages=messages,
                max_completion_tokens=400
            )
            
            return response
        except Exception as e:
            print(f"Error en la solicitud a OpenAI: {e}. Intento {attempt + 1} de {max_retries}.")
            
            # Verificar de nuevo antes de dormir
            if stop_requested_check and stop_requested_check():
                print("Solicitud a OpenAI cancelada durante reintento")
                return None
                
            if attempt < max_retries - 1:
                time.sleep(10)
            else:
                raise

# Modificar assign_labels_to_response para pasar stop_requested
def assign_labels_to_response(question, response, labels, codes, is_single_response=False, stop_requested_check=None):
    labels_str = ', '.join([f"{label} (code: {code})" for label, code in zip(labels, codes)])
    messages = [
        {"role": "system", 
         "content": "You are an expert in coding survey responses with a focus on both 'thematic match' and 'conceptual match.' Assign codes accurately, concisely, and strictly based on the provided instructions without additional comments."},
        {"role": "user", 
         "content": f"""The question is: {question}
         The response is: {response}
         The available codes are: {labels_str}
         VERY IMPORTANT:
         ALWAYS check existing labels first and reuse them if they match conceptually and thematic match.
         Instructions:
         Si hay varios codigos repetidos o con ideas similares que se puedan usar en una respuesta solo usa uno no los uses todos en una respuesta, usa el que más se ajuste textualmente.
         Si la respuesta no es coherente a la pregunta asigna 99
         If no existing code fits the response, reply with 'NEW_LABEL_NEEDED' instead of assigning any code.
         1. Assign only the codes that fit the response based on thematic or conceptual alignment.Not use codes 66, 77, 88, and 99.
         2. Only if in {labels_str} there are no labels or codes other than codes 66, 77, 88, 99 respond with 'NEW_LABEL_NEEDED'
         3. Be conservative in code assignment - it's better to assign fewer, highly relevant codes than too many.
         4. Using codes 66, 77, 88, and 99 unless strictly necessary.
         5. Do not combine codes 66, 77, 88, or 99 with other codes or with each other.
         6. Provide only the numeric codes in your response, separated by semicolons if multiple codes are assigned.
         7. Do not assign more than 6 codes per answer.
         9. If the answer is not logical text and is just signs or symbols, assign code 99.
         10. Assign only one code if this is a single response question.
         """}
    ]
    
    # Pasar la función de verificación de stop_requested
    response = request_openai(messages, stop_requested_check=stop_requested_check)
    
    # Si se detuvo el proceso, devolver un valor predeterminado
    if response is None:
        return "77"  # Código predeterminado cuando se detiene el proceso
        
    assigned_codes = response.choices[0].message.content.strip()
    if is_single_response:
        assigned_codes = assigned_codes.split(';')[0].strip()
    return assigned_codes

# Modificar create_new_labels para pasar stop_requested
def create_new_labels(question, response, available_labels, available_codes, codes_df, stop_requested_check=None):
    response_str = str(response).strip().lower()

    # Verificar si ya existe una etiqueta similar entre las etiquetas actuales o recientemente agregadas
    all_existing_labels = list(available_labels) + list(codes_df.loc[codes_df['Nombre de la Pregunta'] == question, 'Label'].str.lower())
    normalized_labels = [normalize_text(label) for label in all_existing_labels]

    normalized_response = normalize_text(response_str)
    if normalized_response in normalized_labels:
        index_in_labels = normalized_labels.index(normalized_response)
        if index_in_labels < len(available_labels):
            # Si la etiqueta ya existe en available_labels, devuélvela
            print(f"Etiqueta existente encontrada para '{response_str}', reutilizando la etiqueta.")
            return available_labels[index_in_labels]
    
    # Definir el mensaje para la API de OpenAI
    
    messages = [
        {"role": "system", "content": """You are an expert in coding survey responses.
        Your task is to either reuse an existing label or create a new one that can be reused for similar responses.
        
        Antes de generar una etiqueta asegurate de que esa idea se repite al menos 3 veces en mas respuestas. 
        
        VERY IMPORTANT:
        Recuerda que las etiquetas creadas deben estar en español latinoamericano (Colombiano)
        The new labels created must be with perfect spelling.
        ALWAYS check existing labels first and reuse them if they match conceptually
        
        If the answer is a name of a television show, series or movie, first make sure that the television show, series or movie actually exists and second, if you must create the tag, create it with the name correctly written as the response may have errors spelling 

        Don't forget to first check the existing tags so as not to create repeated tags written in different ways, for example "Betty la Fea" and "bety la fea"
        
        Rules for label creation/selection:
        1. ALWAYS check existing labels first and reuse them if they match conceptually. Do not use codes 66, 77, 777, 88, 888, 99, 999.
        2. Labels must be short (4-6 words maximum).
        3. Use general categories that can be applied to multiple responses.
        4. Standardize similar concepts under a single label.
        5. Return only the label, no explanations.
        6. Do not create an "Otro" or "Otros" tag.
        """},
        {"role": "user", "content": f"""
        Question: {question}
        Response to code: {response_str}
        
        Current available labels: {available_labels}
        
        Instructions:
        1. The label must have excellent spelling.
        2. Check if the response matches ANY existing label conceptually.
        3. If yes, return that label.
        4. If no, create a new general label that can be reused.
        5. Return only the label text, no explanations. 
        6. No response can be left without an assigned code.
        """}
    ]

    
    response = request_openai(messages, stop_requested_check=stop_requested_check)
    
    # Si se detuvo el proceso, devolver None
    if response is None:
        return None
        
    result = response.choices[0].message.content.strip()

    # return result
    #Validar el formato de la respuesta y evitar etiquetas duplicadas
    if result and not any(char in result for char in "()") and normalize_text(result) not in normalized_labels:
        print(f"Nueva etiqueta generada: {result}")
        return result
    else:
        print(f"Advertencia: Formato inesperado o etiqueta duplicada para la nueva etiqueta: {result}")
        return None


def normalize_text(text):
    text = text.lower()
    text = re.sub(r'[^\w\s]', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def filter_exclusive_codes(assigned_codes_list):
    exclusive_codes = {'66', '77', '88', '99', '777', '888', '999'}
    non_exclusive_codes = [code for code in assigned_codes_list if code not in exclusive_codes]
    
    if non_exclusive_codes:
        return non_exclusive_codes
    elif assigned_codes_list:
        return [assigned_codes_list[0]]  # Retorna solo el primer código excluyente si solo hay códigos excluyentes
    else:
        return []

def get_next_valid_code(existing_codes):
    """
    Obtiene el siguiente código válido, excluyendo los códigos reservados.
    """
    valid_codes = [int(code) for code in existing_codes if int(code) not in {66, 77, 88, 99, 777, 888, 999}]
    next_code = max(valid_codes, default=0) + 1
    while next_code in {66, 77, 88, 99, 777, 888, 999}:
        next_code += 1
    return f"{next_code:02}"  # Aseguramos que sea de dos dígitos

# Modificar process_response para pasar stop_requested
def process_response(question, response, available_labels, available_codes, limit_77, limit_labels, codes_df, stop_requested_check=None):
    response_str = str(response).strip().lower()
    is_single_response = '(respuesta única)' in question

    # Filtrar códigos excluyentes solo para la pregunta actual
    excluded_codes = {'66', '77', '88', '99', '777', '888', '999'}
    filtered_labels_codes = [
        (label, code) for label, code in zip(available_labels, available_codes)
        if code not in excluded_codes
    ]
    filtered_labels, filtered_codes = zip(*filtered_labels_codes) if filtered_labels_codes else ([], [])

    print(f"[Processing response for question '{question}']")

    # Intentar asignar una etiqueta existente
    assigned_codes = assign_labels_to_response(
        question, response_str, filtered_labels, filtered_codes, 
        is_single_response, stop_requested_check
    )
    
    # Si no hay un código asignado y se necesita una nueva etiqueta
    if assigned_codes == "NEW_LABEL_NEEDED" or assigned_codes == "":
        print(f"Etiqueta nueva necesaria para la respuesta: '{response_str}'")

        # Verificar si se ha alcanzado el límite de etiquetas nuevas
        if limit_labels['count'] >= limit_labels['max']:
            print(f"Límite de nuevas etiquetas alcanzado para la pregunta '{question}'. Asignando código 77.")
            assigned_codes = "77"  # Código por defecto cuando no se pueden crear más etiquetas
        else:
            # Crear una nueva etiqueta usando create_new_labels
            new_label = create_new_labels(
                question, response_str, filtered_labels, filtered_codes, 
                codes_df, stop_requested_check
            )
            
            if new_label:
                # Determinar el nuevo código utilizando get_next_valid_code
                existing_codes = codes_df.loc[codes_df['Nombre de la Pregunta'] == question, 'Cod']
                new_code = get_next_valid_code(existing_codes)
                codes_df, label_created = save_new_label(codes_df, question, new_label, new_code)
                
                if label_created:
                    print(f"Nueva etiqueta creada: '{new_label}' con código {new_code}")

                    # Actualizar códigos y etiquetas disponibles
                    available_codes.append(new_code)
                    available_labels.append(new_label)

                    # Incrementar el contador de etiquetas nuevas
                    limit_labels['count'] += 1
                    limit_77['new_labels'].append((question, new_label, new_code))

                    # Actualizar preguntas en `questions_dict`
                    if question in questions_dict:
                        questions_dict[question].add((new_code, new_label))
                    else:
                        questions_dict[question] = {(new_code, new_label)}
                    
                    assigned_codes = new_code
                else:
                    print(f"No se pudo crear una nueva etiqueta para '{response_str}', asignando código 77")
                    assigned_codes = "77"
            else:
                print(f"No se generó una nueva etiqueta para '{response_str}', asignando código 77")
                assigned_codes = "77"
    else:
        print(f"Etiqueta existente asignada: '{assigned_codes}' para la respuesta '{response_str}'")

    # Convertir los códigos asignados en una lista y eliminar duplicados
    assigned_codes_list = re.findall(r'\d+', str(assigned_codes))
    assigned_codes_list = [f"{int(code):02d}" for code in assigned_codes_list]
    assigned_codes_list = list(set(assigned_codes_list))

    # Si es de respuesta única, solo permitimos un código
    if is_single_response:
        assigned_codes_list = assigned_codes_list[:1]
        
    final_codes = ';'.join(assigned_codes_list)

    print(f"Códigos asignados finales para la respuesta '{response_str}': {final_codes}")
    return final_codes, codes_df

def group_labels_codes(selected_questions, response_columns):
    questions_dict = {}

    for _, row in selected_questions.iterrows():
        if pd.isna(row['Id campo']):
            continue
        
        id_campos = [campo.strip() for campo in row['Id campo'].split('-')]
        
        for col in response_columns:
            col_base = col[:-5] if col.endswith('_OTRO') or col.endswith('_OTRA') else f"C{col}"
            if any(col_base == id_campo for id_campo in id_campos):
            #if col_base in row['Id campo']:
                questions = row['Nombre de la Pregunta'].split(' / ')
                labels = str(row['Label']).split(',')
                codes = str(row['Cod']).split(',') if not pd.isna(row['Cod']) else []

                for question in questions:
                    if question not in questions_dict:
                        questions_dict[question] = set()
                    questions_dict[question].update(zip(codes, labels))
    
    print(f"questions_dict {questions_dict}")
    
    return questions_dict

# Actualizar process_responses para pasar stop_requested
def process_responses(responses_df, codes_df, response_columns, question_column, limit_77, limit_labels, status_text, progress_bar, page):
    global PROCESS_STOPPED
    
    print("EJECUTANDO PROCESS_RESPONSES")
    selected_questions = select_columns(codes_df, question_column)
    global questions_dict
    questions_dict = group_labels_codes(selected_questions, response_columns)
    new_labels = []

    # Crear una copia de codes_df para mantener los cambios
    updated_codes_df = codes_df.copy()
    
    # Calcular total de registros
    total_records = sum(len(responses_df[col].dropna().unique()) for col in response_columns)
    processed_records = 0

    # Modificar esta función para usar la variable global PROCESS_STOPPED
    def check_stop():
        global PROCESS_STOPPED
        return PROCESS_STOPPED

    # Crear un mapeo de columnas a preguntas relevantes
    column_to_questions = {}
    for _, row in selected_questions.iterrows():
        if pd.isna(row['Id campo']):
            continue
        
        id_campos = [campo.strip() for campo in row['Id campo'].split('-')]
        questions = row['Nombre de la Pregunta'].split(' / ')
        
        for col in response_columns:
            col_base = col[:-5] if col.endswith('_OTRO') or col.endswith('_OTRA') else f"C{col}"
            if any(col_base == id_campo for id_campo in id_campos):
                if col not in column_to_questions:
                    column_to_questions[col] = set()
                column_to_questions[col].update(questions)

    for i, col in enumerate(response_columns):
        if PROCESS_STOPPED:
            return responses_df, codes_df  # Retornar los datos procesados hasta ahora
            
        status_text.value = f"Procesando columna {i + 1} de {len(response_columns)}: {col}"
        page.update()
        
        # Obtener solo las preguntas relevantes para esta columna
        relevant_questions = column_to_questions.get(col, set())
        if not relevant_questions:
            print(f"⚠️ No se encontraron preguntas asociadas a la columna {col}. Omitiendo.")
            continue
            
        print(f"Preguntas relevantes para columna {col}: {relevant_questions}")
        
        if col.endswith('_OTRO') or col.endswith('_OTRA'):
            # Procesar columnas tipo "_OTRO" o "_OTRA"
            responses_df, updated_codes_df = process_other_columns(
                responses_df, [col], questions_dict, updated_codes_df,
                progress_bar, status_text, page, total_records, check_stop
            )
        else:
            # Código existente para columnas normales
            code_column = f'C{col}'
            if code_column not in responses_df.columns:
                responses_df[code_column] = ""

            # Formatear códigos existentes
            responses_df[code_column] = responses_df[code_column].apply(
                lambda x: ';'.join([f"{int(cod):02}" for cod in str(x).split(';') if cod.strip().isdigit()]) if pd.notna(x) else ""
            )

            # Procesar respuestas únicas
            unique_responses = responses_df[col].dropna().unique()
            for i, response in enumerate(unique_responses):
                if PROCESS_STOPPED:
                    break  # Salir del bucle de respuestas
                    
                if i % max(1, len(unique_responses)//100) == 0:
                    status_text.value = f"Procesando {col}: {i+1}/{len(unique_responses)}"
                    page.update()
                
                # CAMBIO IMPORTANTE: Iterar solo sobre preguntas relevantes para esta columna
                for question in relevant_questions:
                    if question not in questions_dict:
                        continue
                        
                    data = questions_dict[question]
                    available_codes, available_labels = zip(*data)
                    available_labels = list(available_labels)
                    available_codes = list(available_codes)

                    # Procesar la respuesta con la pregunta específica
                    assigned_codes, updated_codes_df = process_response(
                        question, response, available_labels, available_codes, 
                        limit_77, limit_labels, updated_codes_df, check_stop
                    )

                    # Actualizar solo cuando tengamos una pregunta válida para esta columna
                    mask = responses_df[col] == response
                    responses_df.loc[mask, code_column] = assigned_codes
                    
                    # Registrar las celdas modificadas
                    global MODIFIED_CELLS
                    modified_indices = responses_df.index[mask].tolist()
                    for idx in modified_indices:
                        MODIFIED_CELLS.add((idx, code_column))
                    
                    # Incrementar contador de progreso
                    processed_records += 1
                    progress_bar.value = processed_records / total_records
                    page.update()
                    
                    # Una vez procesada la primera pregunta relevante, salimos del bucle
                    # ya que solo necesitamos una asignación de código por respuesta
                    break

    return responses_df, updated_codes_df

def update_codes_file(codes_df, new_labels):
    
    excluded_codes = {66, 77, 88, 99, 0, 777, 888, 999}

    for id_campo, label, _ in new_labels:
        clean_label = re.sub(r'\(\d{3}\)', '', label).strip()
        print(f"clean_label: {clean_label}")
        

        # Normalizar los valores para asegurar la coincidencia (remover espacios en blanco y convertir a mayúsculas)
        codes_df['Id campo'] = codes_df['Id campo'].astype(str).str.strip().str.upper()
        id_campo_normalized = str(id_campo).strip().upper()
        print(f"id_campo_normalized: {id_campo_normalized}")
        

        # Buscar las filas correspondientes a la pregunta usando la columna 'Id campo'
        question_rows = codes_df.loc[codes_df['Id campo'] == id_campo_normalized]
        print(f"question_rows: {question_rows}")
        

        if question_rows.empty:
            print(f"Warning: Question with Id campo '{id_campo}' not found in DataFrame. Skipping update.")
            continue

        # Obtener los códigos existentes para la pregunta, excluyendo los códigos no válidos
        existing_codes_question = question_rows['Cod'].astype(int)
        print(f"existing_codes_question: {existing_codes_question}")
        

        valid_codes = [cod for cod in existing_codes_question if cod not in excluded_codes]
        print(f"valid_codes: {valid_codes}")
        

        # Asignar un nuevo código que sea el máximo de los existentes + 1
        new_code = max(valid_codes, default=0) + 1
        print(f"new_code: {new_code}")
        

        # Obtener los valores de 'Id campo' y '# Pregunta del formulario' a partir de las filas encontradas
        form_question = question_rows['# Pregunta del formulario'].ffill().bfill().values[0]
        print(f"form_question: {form_question}")
        

        # Crear una nueva fila con la etiqueta limpia y el nuevo código
        new_row = pd.DataFrame({
            'Id campo': [id_campo],
            'Cod': [f"{new_code:02}"],
            'Label': [clean_label],
            'Agrupación': [None], 
            '# Pregunta del formulario': [form_question], 
            'Nombre de la Pregunta': [None]  # No se asigna porque no todas las filas tienen esta información
        })
        print(f"new_row: {new_row}")

        # Concatenar la nueva fila al DataFrame existente
        codes_df = pd.concat([codes_df, new_row], ignore_index=True)
        print(f"Updated codes_df: {codes_df.tail()}")
        

    return codes_df


def process_other_columns(responses_df, response_columns, questions_dict, update_codes_df, 
                          progress_bar, status_text, page, total_records, stop_requested, start_code=501):
    print("EJECUTANDO PROCESS_OTHER_COLUMNS")
    excluded_codes = {'66', '77', '88', '99', '00', '777', '888', '999'}
    new_labels = []
    processed_records = 0  # Nuevo contador para registros procesados

    for col in response_columns:
        if stop_requested:
            return responses_df, update_codes_df  # Detener y devolver lo procesado
            
        if col.endswith('_OTRO') or col.endswith('_OTRA'):
            col_without_other = col[:-5]
            if col_without_other not in responses_df.columns:
                responses_df[col_without_other] = ""

            # Convertir toda la columna a dos dígitos antes de comenzar
            responses_df[col_without_other] = responses_df[col_without_other].apply(
                lambda x: ';'.join([f"{int(code):02d}" for code in str(x).split(';') if code.strip().isdigit()])
            )
            
            other_responses = responses_df[col]
            # for idx, response in tqdm(other_responses.items(), total=len(other_responses), desc=f"Processing {col}"):
            for idx, response in other_responses.items():
                if pd.isna(response):
                    continue

                for question, data in questions_dict.items():
                    if isinstance(data, set) and all(isinstance(item, tuple) and len(item) == 2 for item in data):
                        available_labels, available_codes = zip(*data)
                        available_labels = list(available_labels)
                        available_codes = list(available_codes)
                    else:
                        continue

                    # Asignamos el código y actualizamos update_codes_df
                    assigned_codes, update_codes_df = process_response(
                        question, response,
                        available_labels, available_codes,
                        {'count': 0, 'max': start_code, 'new_code': 0, 'new_labels': new_labels},
                        {'count': 0, 'max': 8},
                        update_codes_df
                    )

                    # Obtener códigos actuales en el DataFrame de respuestas
                    current_codes = str(responses_df.at[idx, col_without_other]).strip()
                    current_codes_set = set(current_codes.split(';')) if current_codes else set()
                    new_codes_set = set(str(assigned_codes).split(';')) if assigned_codes else set()

                    # Convertir todos los códigos en current_codes_set y new_codes_set a dos dígitos solo si son numéricos
                    new_codes_set = {f"{int(code):02d}" for code in new_codes_set }#if code.isdigit()}
                    current_codes_set = {f"{int(code):02d}" for code in current_codes_set }#if code.isdigit()}
                    

                    # Manejo de códigos excluyentes
                    combined_codes_set = current_codes_set | new_codes_set
                    non_excluded_codes = combined_codes_set - excluded_codes

                    if non_excluded_codes:
                        # Si hay códigos no excluyentes, eliminamos los excluidos
                        combined_codes_set = non_excluded_codes

                    # Reemplazo específico del código 77
                    if '77' in current_codes_set:
                        current_codes_set.remove('77')
                        replacement_codes = new_codes_set - current_codes_set  # Evitar duplicados
                        if replacement_codes:
                            # Todos los códigos de reemplazo dentro de un único par de corchetes
                            replacement_codes_str = f"[{';'.join(sorted(replacement_codes))}]"
                            final_codes = ';'.join(sorted(current_codes_set)) + (f";{replacement_codes_str}" if current_codes_set else replacement_codes_str)
                        else:
                            # Si no hay reemplazos, solo quitamos el 77
                            final_codes = ';'.join(sorted(current_codes_set))
                    else:
                        # Si no hay 77, solo combinar los códigos evitando duplicados
                        final_codes = ';'.join(sorted(combined_codes_set))

                    # Filtrar nuevamente excluidos si hay más códigos
                    final_codes_list = final_codes.split(';')
                    final_filtered_codes = [code for code in final_codes_list if code not in excluded_codes] or final_codes_list

                    # Guardar el resultado final en el DataFrame de respuestas
                    responses_df.at[idx, col_without_other] = ';'.join(final_filtered_codes)
                    
                    # Registrar la celda modificada
                    global MODIFIED_CELLS
                    MODIFIED_CELLS.add((idx, col_without_other))

                    # Actualizar progreso
                    processed_records += 1
                    progress_bar.value = processed_records / total_records
                    status_text.value = f"Procesando registro {processed_records} de {total_records}"
                    page.update()

    print(f"Nuevas etiquetas creadas: {new_labels}")
    print(f"update_codes_df antes del return en otros: {update_codes_df}")

    return responses_df, update_codes_df


def update_used_columns(original_responses_df, modified_responses_df, modified_columns, save_path):
    for column in modified_columns:
        # Solo actualizar la columna de códigos
        code_column = f'C{column}'
        if code_column in modified_responses_df.columns:
            # Si la columna de códigos no existe en el DataFrame original, crearla
            if code_column not in original_responses_df.columns:
                original_responses_df[code_column] = ""
            
            # Actualizar solo los valores de la columna de códigos
            original_responses_df[code_column] = modified_responses_df[code_column]
            
            print(f"Verificando columna {code_column}...")
            print(f"Registros vacíos encontrados: {original_responses_df[code_column].isna().sum()}")
            print(f"Registros con string vacío: {(original_responses_df[code_column].astype(str) == '').sum()}")
    
    # Guardar el archivo
    original_responses_df.to_excel(save_path, index=False)
    print(f"Updated file saved at {save_path}")

def save_new_label(codes_df, question, label, new_code):
    question_row = codes_df.loc[codes_df['Nombre de la Pregunta'] == question]
    print(f"question_row: {question_row}")
    

    if question_row.empty:
        print(f"Warning: Question '{question}' not found in DataFrame. Skipping update.")
        return codes_df, False  # No se realizó ningún cambio

    id_campo = question_row['Id campo'].ffill().bfill().values[0]
    print(f"id_campo: {id_campo}")
    

    form_question = question_row['# Pregunta del formulario'].ffill().bfill().values[0]
    print(f"form_question: {form_question}")
    

    new_row = pd.DataFrame({
        'Id campo': [id_campo],
        'Cod': [new_code],
        'Label': [label],
        'Agrupación': [None], 
        '# Pregunta del formulario': [form_question], 
        'Nombre de la Pregunta': [question]
    })
    print(f"new_row: {new_row}")
    
    
    codes_df = pd.concat([codes_df, new_row], ignore_index=True)
    print(f"Updated codes_df: {codes_df.tail()}")
    

    return codes_df, True  # Se ha creado una nueva etiqueta

def save_partial_results(responses_df, codes_df, responses_path, codes_path, current_column):
    """
    Guarda resultados parciales después de procesar cada columna.
    
    Args:
        responses_df: DataFrame con las respuestas procesadas hasta el momento
        codes_df: DataFrame con los códigos actualizados
        responses_path: Ruta original del archivo de respuestas
        codes_path: Ruta original del archivo de códigos
        current_column: Columna que acaba de ser procesada
    
    Returns:
        Tuple con las rutas donde se guardaron los archivos
    """
    # Crear carpeta para guardados parciales si no existe
    partial_folder = os.path.join(os.path.dirname(responses_path), "guardados_parciales")
    os.makedirs(partial_folder, exist_ok=True)
    
    # Generar nombres de archivo con timestamp para evitar sobrescrituras
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    
    # Obtener nombres base de los archivos
    responses_basename = os.path.basename(responses_path).replace(".xlsx", "")
    codes_basename = os.path.basename(codes_path).replace(".xlsx", "")
    
    # Crear rutas para los archivos parciales
    partial_responses_path = os.path.join(
        partial_folder, 
        f"{responses_basename}_parcial_col_{current_column}_{timestamp}.xlsx"
    )
    partial_codes_path = os.path.join(
        partial_folder, 
        f"{codes_basename}_parcial_col_{current_column}_{timestamp}.xlsx"
    )
    
    # Guardar los archivos
    responses_df.to_excel(partial_responses_path, index=False)
    codes_df.to_excel(partial_codes_path, index=False)
    
    print(f"✓ Guardado parcial realizado después de procesar columna {current_column}")
    print(f"  - Respuestas: {partial_responses_path}")
    print(f"  - Códigos: {partial_codes_path}")
    
    return partial_responses_path, partial_codes_path







def save_files_automatically(responses_df, codes_df, original_responses_path, original_codes_path, modified_cells=None):
    """
    Guarda automáticamente los archivos procesados con un timestamp y marca las celdas modificadas
    
    Args:
        responses_df: DataFrame de respuestas procesadas
        codes_df: DataFrame de códigos actualizado
        original_responses_path: Ruta original del archivo de respuestas
        original_codes_path: Ruta original del archivo de códigos
        modified_cells: Set de tuplas (fila, columna) de celdas modificadas
        
    Returns:
        tuple: (ruta_respuestas_guardadas, ruta_códigos_guardados)
    """
    import os
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    # Crear timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Obtener directorio base del archivo original de respuestas
    base_dir = os.path.dirname(original_responses_path)
    
    # Crear carpeta de resultados si no existe
    results_dir = os.path.join(base_dir, "resultados_procesados")
    os.makedirs(results_dir, exist_ok=True)
    
    # Generar nombres de archivos con timestamp
    responses_filename = f"respuestas_procesadas_{timestamp}.xlsx"
    codes_filename = f"codigos_actualizados_{timestamp}.xlsx"
    
    # Rutas completas
    responses_save_path = os.path.join(results_dir, responses_filename)
    codes_save_path = os.path.join(results_dir, codes_filename)
    
    # Guardar archivo de respuestas con formato
    wb_responses = Workbook()
    ws_responses = wb_responses.active
    ws_responses.title = "Respuestas"
    
    # Agregar datos al worksheet
    for r in dataframe_to_rows(responses_df, index=False, header=True):
        ws_responses.append(r)
    
    # Crear fill amarillo para celdas modificadas
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Aplicar formato amarillo a las celdas modificadas
    if modified_cells:
        for row_idx, col_name in modified_cells:
            # Encontrar el índice de la columna
            if col_name in responses_df.columns:
                col_idx = list(responses_df.columns).index(col_name) + 1  # +1 porque openpyxl usa índices base 1
                # +2 porque la primera fila son headers y row_idx es base 0
                cell = ws_responses.cell(row=row_idx + 2, column=col_idx)
                cell.fill = yellow_fill
    
    # Guardar archivo de respuestas
    wb_responses.save(responses_save_path)
    
    # Guardar archivo de códigos (sin formato especial por ahora)
    codes_df.to_excel(codes_save_path, index=False)
    
    return responses_save_path, codes_save_path

def stop_process():
    global PROCESS_STOPPED
    PROCESS_STOPPED = True
    status_text.value = "⏳ Deteniendo el proceso... Cancelando todas las solicitudes. Por favor espere."
    stop_button.disabled = True
    page.update()

def main(page: ft.Page):
    # Definir las variables como globales
    global responses_df, codes_df, responses_path, codes_path, status_text, progress_bar
    global columns_input, max_new_labels_input, max_code_77_input, process_button, stop_button
    
    responses_df = None
    codes_df = None
    responses_path = None
    codes_path = None
    
    # Elementos de la interfaz que se usarán en varias funciones
    status_text = ft.Text(value="Estado: Listo", size=14, color="green")
    progress_bar = ft.ProgressBar(width=400, value=0, visible=False)

    
    # Definir estos elementos ahora para que estén disponibles para todas las funciones
    columns_input = ft.TextField(label="Columnas a codificar (separadas por comas)", width=400)
    max_new_labels_input = ft.TextField(label="Límite de nuevas etiquetas", value="25", width=200)
    max_code_77_input = ft.TextField(label="Límite de código 77 (%)", value="3", width=200)
    
    # Definir start_processing antes de usarlo en el botón
    def start_processing(e):
        global responses_df, codes_df, PROCESS_STOPPED, MODIFIED_CELLS
        PROCESS_STOPPED = False
        MODIFIED_CELLS.clear()  # Limpiar el conjunto de celdas modificadas
        
        if not responses_path or not codes_path:
            status_text.value = "❌ Por favor, carga ambos archivos antes de continuar."
            page.update()
            return
        
        # Obtener valores de los campos de entrada
        response_columns = [col.strip() for col in columns_input.value.split(",") if col.strip()]
        try:
            max_new_labels = int(max_new_labels_input.value)
            max_code_77_percent = float(max_code_77_input.value)
        except ValueError:
            status_text.value = "❌ Por favor, ingresa valores válidos para los límites."
            page.update()
            return
            
        # Verificar que se ingresaron columnas
        if not response_columns:
            status_text.value = "❌ Por favor, ingresa al menos una columna a codificar."
            page.update()
            return
        
        # Mostrar el botón de detener y ocultar el de iniciar
        stop_button.visible = True
        stop_button.disabled = False
        process_button.visible = False
        progress_bar.visible = True
        page.update()
        
        try:
            # Cargar archivos
            global responses_df, codes_df
            responses_df, codes_df = load_files(responses_path, codes_path)
            
            # Verificar columnas
            missing_columns = [col for col in response_columns if col not in responses_df.columns]
            if missing_columns:
                status_text.value = f"🚫 Las siguientes columnas no existen: {', '.join(missing_columns)}"
                stop_button.visible = False
                process_button.visible = True
                progress_bar.visible = False
                page.update()
                return
            
            # Configuración de parámetros
            question_column = "Nombre de la Pregunta"
            limit_77 = {
                'count': 0,
                'max': int(len(responses_df) * (max_code_77_percent / 100)),
                'new_code': codes_df['Cod'].astype(int).max() + 1,
                'new_labels': []
            }
            limit_labels = {'count': 0, 'max': max_new_labels}
            
            # Procesar respuestas
            updated_responses_df, updated_codes_df = process_responses(
                responses_df, codes_df, response_columns, question_column, 
                limit_77, limit_labels, status_text, progress_bar, page
            )
            
            # Guardar resultados
            responses_save_path, codes_save_path = save_files_automatically(
                updated_responses_df, updated_codes_df, responses_path, codes_path, MODIFIED_CELLS
            )
            
            # Mostrar mensaje de éxito o interrupción
            if PROCESS_STOPPED:
                status_text.value = "🛑 Proceso detenido. Se guardaron resultados parciales."
            else:
                status_text.value = f"✅ Proceso completado.\n📂 Archivos guardados en:\n{responses_save_path}\n{codes_save_path}"
                
        except Exception as ex:
            status_text.value = f"❌ Error: {str(ex)}"
        finally:
            # Restaurar la interfaz
            stop_button.visible = False
            process_button.visible = True
            progress_bar.visible = False
            page.update()
    process_button = ft.ElevatedButton("Iniciar procesamiento", on_click=start_processing, disabled=True)
    stop_button = ft.ElevatedButton(
        "⛔ Detener proceso", 
        on_click=lambda _: stop_process(),
        color="error",
        visible=False
    )
    

    
    # Funciones auxiliares para la aplicación principal


    
    # Resto de funciones auxiliares como save_files_automatically, etc.
    
    # Ahora define build_main_app con acceso a start_processing
    def build_main_app():
        # Limpiar página antes de construir la nueva interfaz
        page.clean()
        
        # Variables globales para almacenar los DataFrames
        global responses_df, codes_df, responses_path, codes_path
        responses_df = None
        codes_df = None
        responses_path = None
        codes_path = None
        
        # Título de la aplicación
        welcome_text = ft.Text(
            "CNC Coding Helper",
            size=24,
            weight=ft.FontWeight.BOLD,
            color="blue700"
        )
        
        # Elementos de la interfaz
        global status_text, progress_bar
        status_text = ft.Text(value="Carga los archivos para comenzar.", size=16)
        progress_bar = ft.ProgressBar(width=400, value=0, visible=False)
        
        # File pickers
        responses_file_picker = ft.FilePicker(on_result=lambda e: on_file_pick(e, "responses"))
        codes_file_picker = ft.FilePicker(on_result=lambda e: on_file_pick(e, "codes"))
        page.overlay.append(responses_file_picker)
        page.overlay.append(codes_file_picker)
        
        # Definir on_file_pick AQUÍ dentro de build_main_app
        def on_file_pick(e, file_type):
            global responses_path, codes_path  # Usar nonlocal en lugar de global
            if e.files:
                file_path = e.files[0].path
                if file_type == "responses":
                    responses_path = file_path
                    status_text.value = f"✅📃 Archivo de respuestas cargado: {file_path}"
                elif file_type == "codes":
                    codes_path = file_path
                    status_text.value = f"✅📃 Archivo de códigos cargado: {file_path}"
                page.update()
        
        # Método para actualizar estado del botón de proceso
        def update_process_button():
            process_button.disabled = not (responses_path and codes_path)
            page.update()
        
        # Los campos ya están definidos globalmente, no necesitamos redefinirlos
        
        # Botones
        upload_responses_button = ft.ElevatedButton(
            "📃 Cargar archivo de respuestas", on_click=lambda _: responses_file_picker.pick_files()
        )
        upload_codes_button = ft.ElevatedButton(
            "📃 Cargar archivo de códigos", on_click=lambda _: codes_file_picker.pick_files()
        )
        process_button = ft.ElevatedButton("Iniciar procesamiento", on_click=start_processing, disabled=True)
        stop_button = ft.ElevatedButton(
            "⛔ Detener proceso", 
            on_click=lambda _: stop_process(),
            color="error",
            visible=False
        )

        
        # Elementos específicos para administradores

        
        # Método para actualizar estado del botón de proceso
        def update_process_button():
            process_button.disabled = not (responses_path and codes_path)
            page.update()
        
        # Configurar callbacks para file pickers
        responses_file_picker.on_result = lambda e: (on_file_pick(e, "responses"), update_process_button())
        codes_file_picker.on_result = lambda e: (on_file_pick(e, "codes"), update_process_button())
        
        # Construir interfaz completa
        page.add(
            ft.Column(  # Falta esta apertura de ft.Column y su corchete
                [
                    ft.Row(
                        [
                            ft.Text("CNC Coding Helper", size=24, weight=ft.FontWeight.BOLD),
                            welcome_text,
                        ],
                        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                        width=page.width - 40
                    ),
                    status_text,

                    upload_responses_button,
                    upload_codes_button,
                    progress_bar,
                    columns_input,
                    ft.Row([max_new_labels_input, max_code_77_input], alignment=ft.MainAxisAlignment.CENTER),
                    process_button,
                    stop_button,

                ],
                alignment=ft.MainAxisAlignment.CENTER,
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
            )
        )
        
    # Iniciar directamente con la aplicación principal
    build_main_app()



# Ejecutar la aplicación
ft.app(target=main)