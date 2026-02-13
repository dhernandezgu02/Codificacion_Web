"""
Reviewer Logic - Adapted from reviewer-comparator/reviewer-comparator.py
"""
import pandas as pd
import os
import sys
from typing import List, Dict, Any, Callable, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openai import OpenAI

try:
    import config
    from config import openai_api_key_Codifiacion
except ImportError:
    # Try local import if running directly
    try:
        from backend import config
        from backend.config import openai_api_key_Codifiacion
    except ImportError:
         print("Warning: Could not import config")
         openai_api_key_Codifiacion = None

def clean_codes(codes):
    if pd.isna(codes):
        return codes
    codes = str(codes).replace('[', '').replace(']', '').replace("'", "").split(';')
    unique_codes = list(dict.fromkeys(codes))
    priority_codes = ['99', '88', '77', '66']
    filtered_codes = [code for code in unique_codes if code not in priority_codes or len(unique_codes) == 1]
    
    # Formateamos los códigos a dos dígitos
    formatted_codes = ['{:02d}'.format(int(code.strip())) for code in filtered_codes if code.strip().isdigit()]
    return ';'.join(formatted_codes)

def verify_codes_with_openai(question_text, response_text, assigned_codes, valid_codes, valid_labels):
    prompt = (
        f"Dada la siguiente pregunta: '{question_text}', la respuesta: '{response_text}', "
        f"y los códigos asignados: {assigned_codes}. "
        f"Los códigos válidos para esta pregunta son: {valid_codes}, con las siguientes etiquetas correspondientes: {valid_labels}. "
        "Es muy importante que se asignen los códigos que capturen la idea textual de la respuesta. "
        "Lee muy bien la pregunta y la respuesta para asegurarte de que las asignaciones sean correctas. "
        "Si hay errores en los códigos asignados o faltan códigos necesarios, corrige la asignación en formato lista separada por ';'. "
        "Si la asignación es correcta, devuelve la misma lista sin cambios. "
        "Si una idea en la respuesta puede corresponder a múltiples códigos, asigna solo 1 código por idea."
        "Recuerda que los códigos deben estar a dos dígitos y separados por punto y coma. "
    )
    
    # Using the client from openai>=1.0.0 syntax if available, but the original script used old syntax.
    # The project requirements.txt says openai>=1.55.0, so we must use the new client syntax.
    # The original script used `openai.ChatCompletion.create` which is old syntax, but we should use the new `client.chat.completions.create`.
    
    # from openai import OpenAI
    client = OpenAI(api_key=openai_api_key_Codifiacion)
    
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {"role": "system", "content": "Eres un experto en codificación. TU RESPUESTA DEBE SER ÚNICAMENTE LOS CÓDIGOS SEPARADOS POR PUNTO Y COMA (Ej: 01;05). NO ESCRIBAS NADA DE TEXTO ADICIONAL, NI EXPLICACIONES, NI SALUDOS, NI COMILLAS. SOLO NÚMEROS Y ;."},
            {"role": "user", "content": prompt}
        ]
    )
    
    print("\n[OpenAI Reviewer] Solicitud exitosa")
    print("="*50)
    print(f"[OpenAI Reviewer] Response Object: {response}")
    print(f"[OpenAI Reviewer] Content: {response.choices[0].message.content}")
    print("="*50)
    
    corrected_codes = response.choices[0].message.content.strip()
    
    # Validación y limpieza de la respuesta
    digits = [code.strip() for code in corrected_codes.split(';') if code.strip().isdigit()]
    
    # Eliminar duplicados manteniendo el orden
    seen = set()
    unique_digits = []
    for code in digits:
        if code not in seen:
            seen.add(code)
            unique_digits.append(code)
    
    # Formateamos los códigos a dos dígitos
    formatted_codes = ['{:02d}'.format(int(code.strip())) for code in unique_digits]
    return ';'.join(formatted_codes)

def highlight_changes(file_path, modified_cells):
    wb = load_workbook(file_path)
    ws = wb.active
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for cell in modified_cells:
        ws[cell].fill = yellow_fill
    wb.save(file_path)

def col_index_to_letter(col_index):
    letter = ''
    while col_index >= 0:
        letter = chr(ord('A') + (col_index % 26)) + letter
        col_index = col_index // 26 - 1
    return letter

class SurveyReviewer:
    def __init__(self, responses_path: str, codes_path: str, columns_to_check: List[str]):
        self.responses_path = responses_path
        self.codes_path = codes_path
        self.columns_to_check = columns_to_check
        self.progress_callback: Optional[Callable[[float], None]] = None
        self.status_callback: Optional[Callable[[str], None]] = None
        self.stop_flag = False

    def set_progress_callback(self, callback: Callable[[float], None]):
        self.progress_callback = callback

    def set_status_callback(self, callback: Callable[[str], None]):
        self.status_callback = callback

    def stop(self):
        self.stop_flag = True

    def run(self) -> Dict[str, Any]:
        original_responses_df = pd.read_excel(self.responses_path)
        modified_responses_df = original_responses_df.copy()
        codes_df = pd.read_excel(self.codes_path, sheet_name="Codificación")

        total_rows = 0
        for response_column in self.columns_to_check:
            code_column = "C" + response_column
            if response_column in modified_responses_df.columns and code_column in modified_responses_df.columns:
                total_rows += len(modified_responses_df)

        processed_rows = 0
        corrections_made = 0
        modified_cells = []

        if self.status_callback:
            self.status_callback("Iniciando revisión de asignaciones...")

        for response_column in self.columns_to_check:
            if self.stop_flag:
                break
                
            code_column = "C" + response_column
            if response_column in modified_responses_df.columns and code_column in modified_responses_df.columns:
                if self.status_callback:
                    self.status_callback(f"Revisando columna: {response_column} -> {code_column}")
                
                # Find valid codes/questions in codes_df
                # The logic looks for partial match in 'Id campo'
                valid_codes_df = codes_df[codes_df["Id campo"].astype(str).str.contains(code_column, na=False)]
                
                if not valid_codes_df.empty:
                    # Validate consistency (from original script)
                    try:
                        id_fields = str(valid_codes_df.iloc[0]["Id campo"]).split('-')
                        question_texts = str(valid_codes_df.iloc[0]["Nombre de la Pregunta"]).split('/')
                        # Only warning in logs if mismatch, but proceeding if valid_codes_df is not empty
                    except:
                        pass

                    question_text = valid_codes_df.iloc[0]["Nombre de la Pregunta"]
                    valid_codes = valid_codes_df["Cod"].astype(str).tolist()
                    valid_labels = valid_codes_df["Label"].astype(str).tolist()
                else:
                    print(f"Error: No se encontró la pregunta para {code_column} en el libro de códigos.")
                    continue

                # Clean codes first
                modified_responses_df[code_column] = modified_responses_df[code_column].astype(str).apply(clean_codes)
                
                for idx in range(len(modified_responses_df)):
                    if self.stop_flag:
                        break
                        
                    row = modified_responses_df.iloc[idx]
                    response_text = row[response_column]
                    assigned_codes = row[code_column]
                    
                    if pd.isna(response_text) or str(response_text).strip() == "":
                        processed_rows += 1
                        continue
                        
                    corrected_codes = verify_codes_with_openai(question_text, response_text, assigned_codes, valid_codes, valid_labels)
                    
                    # Clean and format again
                    formatted_corrected_codes = ';'.join(['{:02d}'.format(int(code.strip())) for code in corrected_codes.split(';') if code.strip().isdigit()])
                    
                    if formatted_corrected_codes != assigned_codes:
                        modified_responses_df.at[idx, code_column] = formatted_corrected_codes
                        corrections_made += 1
                        # Track modified cell for highlighting
                        col_idx = modified_responses_df.columns.get_loc(code_column)
                        modified_cells.append(f"{col_index_to_letter(col_idx)}{idx + 2}")
                    
                    processed_rows += 1
                    
                    if self.progress_callback and total_rows > 0:
                        self.progress_callback(processed_rows / total_rows)

        # Save reviewed file
        save_path = self.responses_path.replace(".xlsx", "_reviewed.xlsx")
        modified_responses_df.to_excel(save_path, index=False)
        highlight_changes(save_path, modified_cells)
        
        return {
            "output_file": save_path,
            "corrections_made": corrections_made,
            "total_reviewed": processed_rows
        }
