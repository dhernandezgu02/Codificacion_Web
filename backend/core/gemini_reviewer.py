"""
Reviewer Logic - Gemini Version
"""
import pandas as pd
import os
import sys
from typing import List, Dict, Any, Callable, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Import Gemini Client
from core.gemini_client import request_gemini
# Import stop flag from logic to share global stop state
from core import gemini_logic as logic

def clean_codes(codes):
    if pd.isna(codes):
        return codes
    codes = str(codes).replace('[', '').replace(']', '').replace("'", "").split(';')
    unique_codes = list(dict.fromkeys(codes))
    priority_codes = ['99', '88', '77', '66']
    filtered_codes = [code for code in unique_codes if code not in priority_codes or len(unique_codes) == 1]
    
    # Formateamos los códigos a dos dígitos
    formatted_codes = ['{:02d}'.format(int(code.strip())) for code in filtered_codes if code.strip().isdigit()]
    return ';'.join(formatted_codes)

def verify_codes_with_gemini(question_text, response_text, assigned_codes, valid_codes, valid_labels):
    prompt = (
        f"Dada la siguiente pregunta: '{question_text}', la respuesta: '{response_text}', "
        f"y los códigos asignados: {assigned_codes}. "
        f"Los códigos válidos para esta pregunta son: {valid_codes}, con las siguientes etiquetas correspondientes: {valid_labels}. "
        "Es muy importante que se asignen los códigos que capturen la idea textual de la respuesta. "
        "Lee muy bien la pregunta y la respuesta para asegurarte de que las asignaciones sean correctas. "
        "Si hay errores en los códigos asignados o faltan códigos necesarios, corrige la asignación en formato lista separada por ';'. "
        "Si la asignación es correcta, devuelve la misma lista sin cambios. "
        "Si una idea en la respuesta puede corresponder a múltiples códigos, asigna solo 1 código por idea."
        "Recuerda que los códigos deben estar a dos dígitos y separados por punto y coma. "
    )
    
    messages = [
        {"role": "system", "content": "Eres un experto en codificación de respuestas de encuestas. Asigna códigos de forma precisa. TU RESPUESTA DEBE SER ÚNICAMENTE LA LISTA DE CÓDIGOS SEPARADOS POR PUNTO Y COMA (ej: '01;05'). NO ESCRIBAS PALABRAS, SOLO NÚMEROS Y ;. Si la asignación es correcta, devuelve los mismos códigos."},
        {"role": "user", "content": prompt}
    ]
    
    response_text = request_gemini(messages, temperature=0.0)
    
    if not response_text:
        # Fallback to original
        digits = [code.strip() for code in str(assigned_codes).split(';') if code.strip().isdigit()]
        return ';'.join(['{:02d}'.format(int(c)) for c in digits])

    corrected_codes = response_text.strip()
    
    # Formateamos los códigos a dos dígitos
    digits = [code.strip() for code in corrected_codes.split(';') if code.strip().isdigit()]
    
    # If no digits found but raw response has content, assume AI failed to follow format
    if not digits and corrected_codes and not corrected_codes.replace(';','').isdigit():
        fallback_digits = [code.strip() for code in str(assigned_codes).split(';') if code.strip().isdigit()]
        return ';'.join(['{:02d}'.format(int(c)) for c in fallback_digits])

    formatted_codes = ['{:02d}'.format(int(code)) for code in digits]
    return ';'.join(formatted_codes)

def highlight_changes(file_path, modified_cells):
    wb = load_workbook(file_path)
    ws = wb.active
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for cell in modified_cells:
        ws[cell].fill = yellow_fill
    wb.save(file_path)

def col_index_to_letter(col_index):
    letter = ''
    while col_index >= 0:
        letter = chr(ord('A') + (col_index % 26)) + letter
        col_index = col_index // 26 - 1
    return letter

class SurveyReviewer:
    def __init__(self, responses_path: str, codes_path: str, columns_to_check: List[str]):
        self.responses_path = responses_path
        self.codes_path = codes_path
        self.columns_to_check = columns_to_check
        self.progress_callback: Optional[Callable[[float], None]] = None
        self.status_callback: Optional[Callable[[str], None]] = None
        self.stop_flag = False

    def set_progress_callback(self, callback: Callable[[float], None]):
        self.progress_callback = callback

    def set_status_callback(self, callback: Callable[[str], None]):
        self.status_callback = callback

    def stop(self):
        self.stop_flag = True

    def run(self) -> Dict[str, Any]:
        original_responses_df = pd.read_excel(self.responses_path)
        modified_responses_df = original_responses_df.copy()
        codes_df = pd.read_excel(self.codes_path, sheet_name="Codificación")

        # Cache para evitar llamadas repetitivas a la IA
        # Key: (pregunta, respuesta, codigos_asignados) -> Value: codigos_corregidos
        review_cache = {}

        total_rows = 0
        for response_column in self.columns_to_check:
            code_column = "C" + response_column
            if response_column in modified_responses_df.columns and code_column in modified_responses_df.columns:
                total_rows += len(modified_responses_df)

        processed_rows = 0
        corrections_made = 0
        modified_cells = []

        if self.status_callback:
            self.status_callback("Iniciando revisión de asignaciones (Gemini)...")

        for response_column in self.columns_to_check:
            if self.stop_flag or logic.PROCESS_STOPPED:
                break
                
            code_column = "C" + response_column
            if response_column in modified_responses_df.columns and code_column in modified_responses_df.columns:
                if self.status_callback:
                    self.status_callback(f"Revisando columna: {response_column} -> {code_column}")
                
                # Find valid codes/questions in codes_df
                # The logic looks for partial match in 'Id campo'
                valid_codes_df = codes_df[codes_df["Id campo"].astype(str).str.contains(code_column, na=False)]
                
                if not valid_codes_df.empty:
                    try:
                        id_fields = str(valid_codes_df.iloc[0]["Id campo"]).split('-')
                        question_texts = str(valid_codes_df.iloc[0]["Nombre de la Pregunta"]).split('/')
                    except:
                        pass

                    question_text = valid_codes_df.iloc[0]["Nombre de la Pregunta"]
                    valid_codes = valid_codes_df["Cod"].astype(str).tolist()
                    valid_labels = valid_codes_df["Label"].astype(str).tolist()
                else:
                    print(f"Error: No se encontró la pregunta para {code_column} en el libro de códigos.")
                    continue

                # Clean codes first
                modified_responses_df[code_column] = modified_responses_df[code_column].astype(str).apply(clean_codes)
                
                for idx in range(len(modified_responses_df)):
                    if self.stop_flag or logic.PROCESS_STOPPED:
                        break
                        
                    row = modified_responses_df.iloc[idx]
                    response_text = row[response_column]
                    assigned_codes = row[code_column]
                    
                    if pd.isna(response_text) or str(response_text).strip() == "":
                        processed_rows += 1
                        continue
                    
                    # Verificar caché
                    cache_key = (question_text, str(response_text).strip(), str(assigned_codes).strip())
                    if cache_key in review_cache:
                        corrected_codes = review_cache[cache_key]
                        # print(f"Usando resultado en caché para: {str(response_text)[:20]}...")
                    else:
                        corrected_codes = verify_codes_with_gemini(question_text, response_text, assigned_codes, valid_codes, valid_labels)
                        review_cache[cache_key] = corrected_codes
                    
                    # Clean and format again
                    formatted_corrected_codes = ';'.join(['{:02d}'.format(int(code.strip())) for code in corrected_codes.split(';') if code.strip().isdigit()])
                    
                    if formatted_corrected_codes != assigned_codes:
                        modified_responses_df.at[idx, code_column] = formatted_corrected_codes
                        corrections_made += 1
                        # Track modified cell for highlighting
                        col_idx = modified_responses_df.columns.get_loc(code_column)
                        modified_cells.append(f"{col_index_to_letter(col_idx)}{idx + 2}")
                    
                    processed_rows += 1
                    
                    if self.progress_callback and total_rows > 0:
                        self.progress_callback(processed_rows / total_rows)

        # Save reviewed file
        save_path = self.responses_path.replace(".xlsx", "_reviewed.xlsx")
        modified_responses_df.to_excel(save_path, index=False)
        highlight_changes(save_path, modified_cells)
        
        return {
            "output_file": save_path,
            "corrections_made": corrections_made,
            "total_reviewed": processed_rows
        }
